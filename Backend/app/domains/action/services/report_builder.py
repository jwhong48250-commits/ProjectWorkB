# app/domains/action/services/report_builder.py
import io
import json
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import markdown as md_lib

from app.domains.action import repository
from app.domains.action.models import ReportFormat
from app.domains.action.mongo_repository import get_meeting_summary
from app.domains.action.services import thumbnail as thumb
from app.domains.action.services.wbs_builder import build_wbs_template
from app.utils.s3_utils import upload_fileobj_to_s3

def _thumb_key(meeting_id: int, suffix: str) -> str:
    return f"meetings/{meeting_id}/thumb/thumb_{suffix}.webp"

# ── 색상 상수 ─────────────────────────────────────────────────────────────────
_PRIMARY   = "5668F3"  # 헤더 배경
_SECONDARY = "EEF0FE"  # 섹션 배경
_BORDER_C  = "D0D5F5"  # 테두리

def _report_file_key(meeting_id: int, report_id: int, ext: str) -> str:
    return f"meetings/{meeting_id}/reports/{report_id}.{ext}"

# ── Excel 스타일 헬퍼 ─────────────────────────────────────────────────────────
def _thin_border() -> Border:
    '''
    셀 4면에 테두리를 얇은 연보라 테두리로 적용하는 함수

    셀 4면에 동일한 얇은 테두리를 만든다. 
    _BORDER_C = D0D5F5 => 연보라 16진수(Hex) 컬러 코드
    '''
    s = Side(style="thin", color=_BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(ws, row: int, headers: list[str], widths: list[int]) -> None:
    '''
    헤더의 배경을 파란색으로, 폰트를 흰색 등등으로 적용하는 함수

    PatternFill : 셀의 배경색을 칠함 -> 파란색
    FFFFFF : 하얀색
    Alignment: 정렬 방식 설정
        horizontal: 가로 ,vertical: 세로 -> 둘다 center면 정중앙
        wrap_text: 자동 줄바꿈

    각 셀마다 스타일 세팅
    '''
    fill = PatternFill(start_color=_PRIMARY, end_color=_PRIMARY, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11, name="맑은 고딕")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width

def _apply_row(ws, row: int, values: list, alt: bool = False) -> None:
    '''
    본문의 한줄은 흰색 다음줄은 보라색 계열로 적용하는 함수
    -> 한줄 씩 완성시키는 함수
    
    _SECONDARY = EEF0FE -> 보라색 계열
    보라색 계열의 배경을 격줄로 적용
    헤더보다 1pt 작은 맑은 고딕 폰트
    세로 중앙 정렬, 자동 줄바꿈
    행의 높이를 20으로 세팅
    리스트로 전달 받은 데이터대로 차례대로 입력
    '''
    fill = PatternFill(start_color=_SECONDARY, end_color=_SECONDARY, fill_type="solid") if alt else None
    font = Font(size=10, name="맑은 고딕")
    align = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font   = font
        cell.alignment = align
        cell.border = _thin_border()
        if fill:
            cell.fill = fill

def _section_title(ws, row: int, text: str, cols: int) -> None:
    '''
    ws.row_dimensions[row].height = 22
    -> row에 해당하는 행의 높이를 22pt로 지정
    해당 row의 첫번째 컬럼(열)을 text로 지정한다.
    _PRIMARY = 5668F3 -> 파란색
    font 바꾸기 -> 맑은 고딕 
    merge_cells : 셀 병합
    '''
    ws.row_dimensions[row].height = 22
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color=_PRIMARY, name="맑은 고딕")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

# ── HTML 템플릿 ───────────────────────────────────────────────────────────────
_HTML_CSS = """
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Pretendard', 'Apple SD Gothic Neo', sans-serif;
         background: #f8f9ff; color: #1a1a2e; line-height: 1.7; }
  .container { max-width: 820px; margin: 40px auto; padding: 0 24px 60px; }
  .header { background: linear-gradient(135deg, #5668F3, #7c3aed);
            color: white; border-radius: 16px; padding: 36px 40px; margin-bottom: 32px; }
  .header h1 { font-size: 24px; font-weight: 700; margin-bottom: 8px; }
  .header .meta { font-size: 14px; opacity: 0.85; display: flex; gap: 20px; flex-wrap: wrap; }
  .section { background: white; border-radius: 12px; padding: 28px 32px;
             margin-bottom: 20px; box-shadow: 0 1px 4px rgba(86,104,243,0.08); }
  .section-title { font-size: 16px; font-weight: 700; color: #5668F3;
                   margin-bottom: 16px; padding-bottom: 10px;
                   border-bottom: 2px solid #EEF0FE; display: flex; align-items: center; gap: 8px; }
  .section-title::before { content: ''; width: 4px; height: 18px;
                            background: #5668F3; border-radius: 2px; display: inline-block; }
  .badge { display: inline-block; padding: 2px 10px; border-radius: 20px;
           font-size: 12px; font-weight: 600; }
  .badge-high { background: #fee2e2; color: #dc2626; }
  .badge-normal, .badge-medium { background: #e0f2fe; color: #0369a1; }
  .badge-low { background: #f0fdf4; color: #16a34a; }
  .badge-urgent { background: #fef3c7; color: #d97706; }
  table { width: 100%; border-collapse: collapse; font-size: 14px; }
  th { background: #EEF0FE; color: #5668F3; font-weight: 600; padding: 10px 14px;
       text-align: left; border: 1px solid #d0d5f5; }
  td { padding: 10px 14px; border: 1px solid #e5e7eb; vertical-align: top; }
  tr:nth-child(even) td { background: #f9faff; }
  ul { padding-left: 18px; }
  li { margin-bottom: 6px; font-size: 14px; }
  .tag { background: #EEF0FE; color: #5668F3; border-radius: 6px;
         padding: 2px 8px; font-size: 12px; font-weight: 500; }
  .minutes-content h2 { font-size: 16px; font-weight: 700; color: #5668F3; margin: 20px 0 8px; padding-bottom: 6px; border-bottom: 1px solid #EEF0FE; }
  .minutes-content h3 { font-size: 14px; font-weight: 600; color: #1a1a2e; margin: 12px 0 6px; }
  .minutes-content p  { margin-bottom: 8px; }
  .minutes-content ul { padding-left: 18px; margin: 6px 0; }
  .minutes-content li { margin-bottom: 4px; }
</style>
"""

def _build_html(meeting_title: str, summary: dict, minutes_content: str) -> str:
    # summary 에서 각 데이터 꺼내기
    overview  = summary.get("overview", {})
    attendees = summary.get("attendees", [])
    decisions = summary.get("decisions", [])
    actions   = summary.get("action_items", [])
    pending   = summary.get("pending_items", [])

    # 우선순위/ 긴급도를 색깔 뱃지 HTML로 변환
    def badge(value: str, kind: str = "priority") -> str:
        v = (value or "").lower()
        css = f"badge-{v}" if v in ("high", "normal", "medium", "low", "urgent") else "badge-normal"
        return f'<span class="badge {css}">{value}</span>'

    decisions_rows = "".join(
        f"<tr><td>{d.get('decision','')}</td><td>{d.get('rationale','')}</td>"
        f"<td>{d.get('opposing_opinion','') or '-'}</td></tr>"
        for d in decisions
    )
    action_rows = "".join(
        f"<tr><td><strong>{a.get('assignee','')}</strong></td>"
        f"<td>{a.get('content','')}</td>"
        f"<td>{a.get('deadline','') or '-'}</td>"
        f"<td>{badge(a.get('priority',''))}</td>"
        f"<td>{badge(a.get('urgency',''))}</td></tr>"
        for a in actions
    )
    pending_items = "".join(
        ("<li>" + p.get('content','') + (' <span class="tag">이월</span>' if p.get('carried_over') else '') + "</li>")
        for p in pending
    ) or "<li>없음</li>"

    minutes_html = md_lib.markdown(
        minutes_content,
        extensions=["tables", "fenced_code"]
    )

    return f"""<!DOCTYPE html>
            <html lang="ko">
            <head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
            <title>{meeting_title} 회의록</title>{_HTML_CSS}</head>
            <body>
            <div class="container">
            <div class="header">
                <h1>{meeting_title}</h1>
                <div class="meta">
                <span>📅 {overview.get('datetime_str', '')}</span>
                <span>👥 {', '.join(attendees)}</span>
                <span>🎯 {overview.get('purpose', '')}</span>
                </div>
            </div>

            <div class="section">
                <div class="section-title">회의 내용</div>
                <div class="minutes-content">{minutes_html}</div>
            </div>

            {'<div class="section"><div class="section-title">결정사항</div><table><thead><tr><th>결정</th><th>근거</th><th>반대의견</th></tr></thead><tbody>' + decisions_rows + '</tbody></table></div>' if decisions else ''}

            {'<div class="section"><div class="section-title">액션 아이템</div><table><thead><tr><th style="white-space:nowrap;min-width:80px">담당자</th><th>내용</th><th>마감</th><th>우선순위</th><th>긴급도</th></tr></thead><tbody>' + action_rows + '</tbody></table></div>' if actions else ''}

            {'<div class="section"><div class="section-title">미결사항</div><ul>' + pending_items + '</ul></div>' if pending else ''}
            </div>
            </body></html>"""

# ── 생성 함수 ─────────────────────────────────────────────────────────────────
def generate_markdown(db: Session, meeting_id: int, user_id: int):
    minute = repository.get_meeting_minute(db, meeting_id)
    if not minute or not minute.content:
        raise ValueError("회의록이 없습니다.")

    meeting    = repository.get_meeting(db, meeting_id)
    thumb_bytes = thumb.generate_text_thumbnail_bytes(minute.content)
    thumb_key = _thumb_key(meeting_id, "md")
    upload_fileobj_to_s3(
        fileobj=io.BytesIO(thumb_bytes),
        key=thumb_key,
        content_type="image/webp",
    )

    return repository.save_report(
        db, meeting_id, user_id,
        format=ReportFormat.markdown,
        title=f"{meeting.title} 회의록",
        content=minute.content,
        thumbnail_url=thumb_key,
    )


def generate_html(db: Session, meeting_id: int, user_id: int):
    minute = repository.get_meeting_minute(db, meeting_id)
    if not minute or not minute.content:
        raise ValueError("회의록이 없습니다.")

    summary = get_meeting_summary(meeting_id)
    meeting = repository.get_meeting(db, meeting_id)

    thumb_bytes = thumb.generate_text_thumbnail_bytes(minute.content)
    thumb_key = _thumb_key(meeting_id, "html")
    upload_fileobj_to_s3(
        fileobj=io.BytesIO(thumb_bytes),
        key=thumb_key,
        content_type="image/webp",
    )

    # HTML은 다운로드 시 즉시 생성 — content 컬럼에 HTML 저장
    html = _build_html(meeting.title, summary, minute.content)

    return repository.save_report(
        db=db,
        meeting_id=meeting_id,
        created_by=user_id,
        format=ReportFormat.html,
        title=f"{meeting.title} HTML 보고서",
        content=html,
        thumbnail_url=thumb_key,
    )


def generate_excel(db: Session, meeting_id: int, user_id: int):
    summary = get_meeting_summary(meeting_id)
    if not summary:
        raise ValueError("회의 요약 데이터가 없습니다.")

    meeting = repository.get_meeting(db, meeting_id)
    report  = repository.save_report(
        db=db, meeting_id=meeting_id, created_by=user_id,
        format=ReportFormat.excel,
        title=f"{meeting.title} Excel 보고서",
    )

    file_key = _report_file_key(meeting_id, report.id, "xlsx")
    thumb_key = _thumb_key(meeting_id, "excel")

    wb       = openpyxl.Workbook()
    overview = summary.get("overview", {})

    # ── 시트1: 개요 ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "개요"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55

    _section_title(ws, 1, "📋 회의 개요", 2)
    rows = [
        ("목적",    overview.get("purpose", "")),
        ("일시",    overview.get("datetime_str", "")),
        ("참석자",  ", ".join(summary.get("attendees", []))),
        ("다음 회의", summary.get("next_meeting", "")),
    ]
    for i, (label, value) in enumerate(rows, 2):
        ws.row_dimensions[i].height = 20
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font   = Font(bold=True, size=10, name="맑은 고딕")
        label_cell.fill   = PatternFill(start_color=_SECONDARY, end_color=_SECONDARY, fill_type="solid")
        label_cell.border = _thin_border()
        label_cell.alignment = Alignment(vertical="center")
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.font   = Font(size=10, name="맑은 고딕")
        value_cell.border = _thin_border()
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── 시트2: 결정사항 ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("결정사항")
    _section_title(ws2, 1, "✅ 결정사항", 3)
    _apply_header(ws2, 2, ["결정사항", "근거", "반대의견"], [40, 35, 25])
    for i, d in enumerate(summary.get("decisions", []), 3):
        _apply_row(ws2, i, [d.get("decision",""), d.get("rationale",""), d.get("opposing_opinion","")], i % 2 == 0)

    # ── 시트3: 액션아이템 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("액션아이템")
    _section_title(ws3, 1, "🎯 액션아이템", 5)
    _apply_header(ws3, 2, ["담당자", "내용", "마감일", "우선순위", "긴급도"], [14, 44, 14, 12, 12])
    for i, a in enumerate(summary.get("action_items", []), 3):
        _apply_row(ws3, i, [
            a.get("assignee",""), a.get("content",""),
            a.get("deadline",""), a.get("priority",""), a.get("urgency",""),
        ], i % 2 == 0)

    # ── 시트4: 미결사항 ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("미결사항")
    _section_title(ws4, 1, "⏳ 미결사항", 2)
    _apply_header(ws4, 2, ["내용", "이월여부"], [55, 12])
    for i, p in enumerate(summary.get("pending_items", []), 3):
        _apply_row(ws4, i, [p.get("content",""), "O" if p.get("carried_over") else "X"], i % 2 == 0)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)
    upload_fileobj_to_s3(
        fileobj=excel_buf,
        key=file_key,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    thumb_bytes = thumb.generate_format_thumbnail_bytes("excel")
    upload_fileobj_to_s3(
        fileobj=io.BytesIO(thumb_bytes),
        key=thumb_key,
        content_type="image/webp",
    )

    report.file_url = file_key
    report.thumbnail_url = thumb_key
    db.commit()
    db.refresh(report)
    return report


async def generate_wbs(db: Session, meeting_id: int, user_id: int):
    wbs     = await build_wbs_template(db, meeting_id)
    meeting = repository.get_meeting(db, meeting_id)

    thumb_bytes = thumb.generate_format_thumbnail_bytes("wbs")
    thumb_key = _thumb_key(meeting_id, "wbs")
    upload_fileobj_to_s3(
        fileobj=io.BytesIO(thumb_bytes),
        key=thumb_key,
        content_type="image/webp",
    )

    return repository.save_report(
        db, meeting_id, user_id,
        format=ReportFormat.wbs,
        title=f"{meeting.title} WBS",
        content=json.dumps(wbs, ensure_ascii=False),
        thumbnail_url=thumb_key,
    )
